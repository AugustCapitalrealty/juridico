#!/usr/bin/env python3
"""
Converte uma planilha de pesquisa (.xlsx exportado da plataforma) no dump JSON
consumido por testes/executar-testes.js.

    python3 testes/gerar-dump.py pesquisa.xlsx testes/planilha-exemplo.json

O dump contém dados pessoais (CPF, e-mail, telefone) e por isso está no
.gitignore: gere-o localmente e não versione.

Observação: o openpyxl rejeita abas cujo nome contém "/" (ex.: "TST/CNDT").
O script contorna isso reescrevendo os nomes antes da leitura; o Apps Script
não tem essa limitação, então os nomes originais continuam válidos em produção.
"""

import datetime
import json
import os
import re
import shutil
import sys
import tempfile
import zipfile

try:
    import openpyxl
except ImportError:
    sys.exit('Instale a dependência: pip install openpyxl')


def sanear_nomes_de_abas(caminho_xlsx):
    """Reescreve nomes de aba com '/' e devolve o caminho de um arquivo temporário."""
    destino = os.path.join(tempfile.mkdtemp(), 'planilha.xlsx')
    desempacotado = tempfile.mkdtemp()

    with zipfile.ZipFile(caminho_xlsx) as zf:
        zf.extractall(desempacotado)

    workbook_xml = os.path.join(desempacotado, 'xl', 'workbook.xml')
    with open(workbook_xml, encoding='utf-8') as f:
        conteudo = f.read()

    conteudo = re.sub(
        r'name="([^"]*)"',
        lambda m: 'name="%s"' % m.group(1).replace('/', '-'),
        conteudo,
    )

    with open(workbook_xml, 'w', encoding='utf-8') as f:
        f.write(conteudo)

    with zipfile.ZipFile(destino, 'w', zipfile.ZIP_DEFLATED) as zf:
        for raiz, _, arquivos in os.walk(desempacotado):
            for arquivo in arquivos:
                completo = os.path.join(raiz, arquivo)
                zf.write(completo, os.path.relpath(completo, desempacotado))

    shutil.rmtree(desempacotado)
    return destino


def celula(valor):
    if valor is None:
        return ''
    if isinstance(valor, (datetime.datetime, datetime.date)):
        return valor.isoformat()
    return valor


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)

    origem, destino = sys.argv[1], sys.argv[2]
    temporario = sanear_nomes_de_abas(origem)
    workbook = openpyxl.load_workbook(temporario, data_only=True)

    dump = {}
    for nome in workbook.sheetnames:
        aba = workbook[nome]
        dump[nome] = [[celula(c) for c in linha] for linha in aba.iter_rows(values_only=True)]

    with open(destino, 'w', encoding='utf-8') as f:
        json.dump(dump, f, ensure_ascii=False)

    print('%d abas exportadas para %s' % (len(dump), destino))


if __name__ == '__main__':
    main()
